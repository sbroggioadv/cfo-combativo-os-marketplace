#!/usr/bin/env python3
"""
xlsx_parser.py — Parser de planilhas XLSX -> schema canonico (§3).

Estrategia (premortem C2 — degradacao graciosa, import dentro de funcao):
  1. pandas (pip install pandas openpyxl) — robusto
  2. openpyxl puro (pip install openpyxl) — fallback
  3. se nada: erro ESTRUTURADO instruindo pip install

Detecta:
  - Cabecalho deslocado (planilha com titulo/logo nas primeiras linhas).
  - Linhas-lixo (rodape de saldo, totais, linhas vazias).

Reusa a heuristica de mapeamento de colunas do csv_parser (sinonimos).

USO:
    python3 xlsx_parser.py <arquivo.xlsx> [entidade_id] [conta_id] [banco]
"""

from __future__ import annotations

import io
import json
import re
import sys
from pathlib import Path
from typing import Any

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


def _csv_helpers():
    """Importa helpers do csv_parser (mapeamento/normalizacao de valor)."""
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import csv_parser  # type: ignore
    return csv_parser


def _erro_dependencia() -> dict[str, Any]:
    return {
        "ok": False,
        "erro": "dependencia_indisponivel",
        "mensagem": ("Para ler XLSX instale uma das opcoes:\n"
                     "  - pip install pandas openpyxl  (RECOMENDADO)\n"
                     "  - pip install openpyxl"),
        "transacoes": [],
    }


def _matriz_via_pandas(path: str) -> list[list[Any]] | None:
    try:
        import pandas as pd  # type: ignore  (import DENTRO da funcao)
    except ImportError:
        return None
    try:
        df = pd.read_excel(path, header=None, dtype=str)
        return df.fillna("").values.tolist()
    except Exception as exc:  # noqa: BLE001
        print(f"[xlsx_parser] pandas falhou: {exc}", file=sys.stderr)
        return None


def _matriz_via_openpyxl(path: str) -> list[list[Any]] | None:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        matriz: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            matriz.append(["" if c is None else c for c in row])
        wb.close()
        return matriz
    except Exception as exc:  # noqa: BLE001
        print(f"[xlsx_parser] openpyxl falhou: {exc}", file=sys.stderr)
        return None


def parse_xlsx(path: str, *, entidade_id: str | None = None,
               conta_id: str | None = None, banco: str | None = None) -> dict[str, Any]:
    """Le XLSX de extrato/planilha e retorna transacoes no schema canonico."""
    if not Path(path).exists():
        return {"ok": False, "erro": "arquivo_nao_encontrado", "caminho": path, "transacoes": []}

    matriz = _matriz_via_pandas(path)
    metodo = "pandas"
    if matriz is None:
        matriz = _matriz_via_openpyxl(path)
        metodo = "openpyxl"
    if matriz is None:
        return _erro_dependencia()

    cp = _csv_helpers()
    # detecta cabecalho deslocado: 1a linha com >=2 papeis conhecidos
    todos = {syn for syns in cp._SINONIMOS.values() for syn in syns}
    header_idx = 0
    for i, row in enumerate(matriz[:20]):
        cols = [cp._norm(str(c)) for c in row]
        if sum(1 for c in cols if c in todos) >= 2:
            header_idx = i
            break

    header = [str(c) for c in matriz[header_idx]]
    mapa = cp._mapear_colunas(header)
    banco_det = cp._detectar_banco(header, banco)

    if "data" not in mapa or ("valor" not in mapa and not ("credito" in mapa or "debito" in mapa)):
        return {
            "ok": False, "erro": "layout_nao_reconhecido",
            "mensagem": "Nao identifiquei colunas de data e valor. Informe o mapeamento (PA-14).",
            "header_detectado": header, "transacoes": [],
        }

    crus: list[dict] = []
    for row in matriz[header_idx + 1:]:
        row = [str(c) for c in row]
        if len(row) <= mapa["data"]:
            continue
        data = row[mapa["data"]].strip()
        # normaliza datas que vem como "2026-04-05 00:00:00" do pandas
        data = data.split(" ")[0]
        if not re.search(r"\d{2}[/-]\d{2}[/-]\d{2,4}|\d{4}-\d{2}-\d{2}", data):
            continue  # linha-lixo

        if "valor" in mapa and mapa["valor"] < len(row):
            valor = cp._br_to_float(row[mapa["valor"]])
            sinal = "D" if valor < 0 else "C"
        else:
            cred = cp._br_to_float(row[mapa["credito"]]) if "credito" in mapa and mapa["credito"] < len(row) else 0.0
            deb = cp._br_to_float(row[mapa["debito"]]) if "debito" in mapa and mapa["debito"] < len(row) else 0.0
            valor = cred if cred else -deb
            sinal = "C" if cred else "D"
        if valor == 0:
            continue
        desc = row[mapa["descricao"]].strip() if "descricao" in mapa and mapa["descricao"] < len(row) else ""
        crus.append({"data": data, "valor": abs(valor), "sinal": sinal,
                     "descricao": desc, "banco": banco_det})

    if entidade_id and conta_id:
        try:
            sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "lib"))
            from canonical import normalizar  # type: ignore
            txs = [normalizar(c, entidade_id=entidade_id, conta_id=conta_id,
                              origem="xlsx", banco=c.get("banco")).to_dict() for c in crus]
            return {"ok": True, "metodo": metodo, "banco_detectado": banco_det,
                    "qtd": len(txs), "transacoes": txs}
        except Exception as exc:  # noqa: BLE001
            print(f"[xlsx_parser] normalizacao falhou ({exc}); retornando crus", file=sys.stderr)

    return {"ok": True, "metodo": metodo, "banco_detectado": banco_det,
            "qtd": len(crus), "transacoes": crus}


# ---------------------------------------------------------------------------
# Auto-teste — gera XLSX sintetico SE openpyxl/pandas existir; senao valida
# que a degradacao graciosa funciona (erro estruturado, sem crash).
# ---------------------------------------------------------------------------

def _gerar_xlsx_sintetico(path: Path) -> bool:
    """Tenta criar um XLSX com cabecalho deslocado + linha-lixo. Retorna True se conseguiu."""
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        return False
    wb = Workbook()
    ws = wb.active
    ws.append(["EXTRATO SINTETICO - BANCO TESTE"])   # titulo (cabecalho deslocado)
    ws.append([])                                     # linha vazia
    ws.append(["Data", "Histórico", "Valor"])         # cabecalho real na linha 3
    ws.append(["05/04/2026", "Receita Servico", "5000,00"])
    ws.append(["12/04/2026", "Aluguel", "-1200,50"])
    ws.append(["SALDO FINAL", "", "3799,50"])         # linha-lixo (sem data)
    wb.save(str(path))
    return True


def _auto_teste() -> int:
    import tempfile
    print("== AUTO-TESTE xlsx_parser.py (XLSX sintetico) ==")
    falhas = 0
    with tempfile.TemporaryDirectory() as td:
        f = Path(td) / "extrato.xlsx"
        if _gerar_xlsx_sintetico(f):
            r = parse_xlsx(str(f), entidade_id="alfa", conta_id="itau-001")
            assert r["ok"] and r["qtd"] == 2, f"deveria ler 2 tx (cabecalho desloc + lixo ignorado): {r}"
            assert r["transacoes"][0]["sinal"] == "C" and r["transacoes"][1]["sinal"] == "D"
            assert r["transacoes"][1]["valor"] == 1200.50
            print("  [ok] leu XLSX: cabecalho deslocado detectado, linha SALDO ignorada")
        else:
            # openpyxl ausente: valida degradacao graciosa
            r = _erro_dependencia()
            assert not r["ok"] and "pip install" in r["mensagem"]
            print("  [ok] openpyxl ausente -> erro estruturado com pip install, sem crash (C2)")

        # arquivo inexistente sempre degrada
        r_err = parse_xlsx(str(Path(td) / "nao-existe.xlsx"))
        assert not r_err["ok"], "deveria degradar em arquivo inexistente"
        print("  [ok] arquivo inexistente -> erro estruturado")

    print("RESULTADO: xlsx_parser.py PASSOU" if falhas == 0 else f"RESULTADO: {falhas} FALHAS")
    return 0 if falhas == 0 else 1


def _main() -> int:
    if len(sys.argv) < 2:
        print("USO: python3 xlsx_parser.py <arquivo.xlsx> [entidade_id] [conta_id] [banco]", file=sys.stderr)
        return 2
    res = parse_xlsx(sys.argv[1],
                     entidade_id=sys.argv[2] if len(sys.argv) > 2 else None,
                     conta_id=sys.argv[3] if len(sys.argv) > 3 else None,
                     banco=sys.argv[4] if len(sys.argv) > 4 else None)
    print(json.dumps(res, ensure_ascii=False, indent=2))
    return 0 if res.get("ok") else 1


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--auto-teste":
        sys.exit(_auto_teste())
    sys.exit(_main())
